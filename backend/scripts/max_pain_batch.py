"""
max_pain_batch.py

Reads a list of tickers from a JSON config, fetches each one's options chain
(Call/Put OI by strike) with a robust retry loop, and writes a single Excel
workbook: a "Dashboard" summary tab + one detail tab per ticker with live
formulas and charts.

Data source: Yahoo Finance via `yfinance` (free, delayed/EOD OI -- see notes
in fetch_one()). Not all tickers have an options market accessible this way
-- e.g. DAX/Eurex index options do not -- so failures are expected for some
tickers and are reported clearly rather than crashing the batch.
"""

import argparse
import json
import sys
import time
from datetime import datetime, timedelta

try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


class PermanentFetchError(Exception):
    """Raised when the failure is structural (no options market exists for
    this ticker via this data source) rather than transient."""


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def load_config(path):
    with open(path) as f:
        cfg = json.load(f)

    if isinstance(cfg, list):
        cfg = {"tickers": cfg}

    tickers = []
    for t in cfg.get("tickers", []):
        if isinstance(t, str):
            tickers.append({"symbol": t, "yahoo_ticker": t, "expiration": None})
        else:
            symbol = t["symbol"]
            tickers.append({
                "symbol": symbol,
                "yahoo_ticker": t.get("yahoo_ticker", symbol),
                "expiration": t.get("expiration"),
            })

    cfg["tickers"] = tickers
    cfg.setdefault("strike_range_pct", 20.0)
    cfg.setdefault("max_strikes", 30)
    return cfg


# ---------------------------------------------------------------------------
# Fetching -- real data
# ---------------------------------------------------------------------------

def is_valid_data(rows):
    if not rows:
        return False, "empty chain"
    if sum(c + p for _, c, p in rows) <= 0:
        return False, "all OI values are zero (likely a stale/partial response)"
    return True, None


def compute_max_pain(rows):
    best_strike, best_pain = None, None
    for p, _, _ in rows:
        total = 0.0
        for k, c, pu in rows:
            total += c * max(p - k, 0)
            total += pu * max(k - p, 0)
        if best_pain is None or total < best_pain:
            best_pain, best_strike = total, p
    return best_strike


def fetch_one(yahoo_ticker, expiration, strike_range_pct, max_strikes):
    import yfinance as yf

    stock = yf.Ticker(yahoo_ticker)
    expirations = stock.options
    if not expirations:
        raise PermanentFetchError(
            "no options data available via Yahoo Finance / yfinance for this ticker "
            "(common for non-US-listed underlyings, e.g. Eurex-cleared DAX/ODAX)"
        )

    exp = expiration or expirations[0]
    if exp not in expirations:
        raise PermanentFetchError(f"expiration '{exp}' not available. Choices: {list(expirations)}")

    # Fetch company name
    try:
        company_name = stock.info.get('shortName', yahoo_ticker)
    except Exception:
        company_name = yahoo_ticker

    chain = stock.option_chain(exp)
    calls = chain.calls[["strike", "openInterest"]].fillna(0)
    puts = chain.puts[["strike", "openInterest"]].fillna(0)
    call_oi = dict(zip(calls["strike"], calls["openInterest"]))
    put_oi = dict(zip(puts["strike"], puts["openInterest"]))
    strikes = sorted(set(call_oi) | set(put_oi))

    try:
        last_price = stock.fast_info["last_price"]
    except Exception:
        last_price = stock.history(period="1d")["Close"].iloc[-1]

    rows = [(k, int(call_oi.get(k, 0)), int(put_oi.get(k, 0))) for k in strikes]

    lo = last_price * (1 - strike_range_pct / 100)
    hi = last_price * (1 + strike_range_pct / 100)
    rows = [r for r in rows if lo <= r[0] <= hi]
    if len(rows) > max_strikes:
        atm_idx = min(range(len(rows)), key=lambda i: abs(rows[i][0] - last_price))
        half = max_strikes // 2
        start = max(0, atm_idx - half)
        end = min(len(rows), start + max_strikes)
        start = max(0, end - max_strikes)
        rows = rows[start:end]

    return rows, exp, float(last_price), company_name


# ---------------------------------------------------------------------------
# Fetching -- demo/synthetic
# ---------------------------------------------------------------------------

KNOWN_UNSUPPORTED = {"DAX", "^GDAXI", "ODAX"}


def fetch_one_demo(symbol):
    if symbol.upper() in KNOWN_UNSUPPORTED:
        raise PermanentFetchError(
            "no options data available via Yahoo Finance / yfinance for this ticker "
            "(Eurex-cleared index options are not sourced by this data feed)"
        )

    import random
    random.seed(hash(symbol) % (2**32))
    base = {"MRVL": 78.0, "NVDA": 172.0, "AAPL": 231.0}.get(symbol.upper(), 100.0)
    
    # Mock company name mapping
    name_map = {"MRVL": "Marvell Technology", "NVDA": "NVIDIA Corp", "AAPL": "Apple Inc."}
    company_name = name_map.get(symbol.upper(), f"{symbol.upper()} Corporation")

    step = base * 0.025
    strikes = [round(base + (i - 10) * step, 2) for i in range(21)]
    rows = []
    for i, k in enumerate(strikes):
        dist = abs(i - 10)
        c = max(int(random.gauss(6000 - dist * 450, 400)), 50)
        p = max(int(random.gauss(5500 - dist * 400, 400)), 50)
        rows.append((k, c, p))
    days_out = (4 - datetime.now().weekday()) % 7 + 7 
    exp = (datetime.now() + timedelta(days=days_out)).strftime("%Y-%m-%d")
    return rows, exp, base, company_name


# ---------------------------------------------------------------------------
# Retry driver
# ---------------------------------------------------------------------------

def fetch_with_retry(ticker_cfg, strike_range_pct, max_strikes,
                      max_retry_minutes, retry_interval_sec, demo):
    symbol = ticker_cfg["symbol"]
    deadline = time.monotonic() + max_retry_minutes * 60
    attempt = 0
    last_err = None

    while True:
        attempt += 1
        try:
            if demo:
                rows, exp, price, company_name = fetch_one_demo(symbol)
            else:
                rows, exp, price, company_name = fetch_one(
                    ticker_cfg["yahoo_ticker"], ticker_cfg["expiration"],
                    strike_range_pct, max_strikes,
                )
            ok, reason = is_valid_data(rows)
            if ok:
                print(f"[{symbol}] OK on attempt {attempt}: {len(rows)} strikes, expiration {exp}")
                return {
                    "status": "OK", "rows": rows, "expiration": exp, "price": price,
                    "company_name": company_name,
                    "max_pain": compute_max_pain(rows),
                    "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            last_err = reason
        except PermanentFetchError as e:
            print(f"[{symbol}] permanent failure, not retrying: {e}")
            return {"status": "FAILED", "error": str(e),
                    "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        except Exception as e:
            last_err = str(e)

        if time.monotonic() >= deadline:
            print(f"[{symbol}] FAILED after {attempt} attempts ({max_retry_minutes:.0f} min): {last_err}")
            return {"status": "FAILED", "error": last_err,
                    "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

        print(f"[{symbol}] attempt {attempt} failed ({last_err}); retrying in {retry_interval_sec:.0f}s...")
        time.sleep(retry_interval_sec)


# ---------------------------------------------------------------------------
# Scheduling helper
# ---------------------------------------------------------------------------

def sleep_until(target_hhmm, tz_name):
    if ZoneInfo is None:
        print("zoneinfo unavailable (Python < 3.9?) -- skipping --wait-until, running now.", file=sys.stderr)
        return
    tz = ZoneInfo(tz_name)
    now = datetime.now(tz)
    hh, mm = map(int, target_hhmm.split(":"))
    target = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
    if target <= now:
        target += timedelta(days=1)
    wait_sec = (target - now).total_seconds()
    print(f"Waiting until {target.isoformat()} ({wait_sec / 60:.1f} min)...")
    time.sleep(wait_sec)


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
MAXPAIN_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
FAIL_FILL = PatternFill(fill_type="solid", start_color="F8CBAD", end_color="F8CBAD")
OK_FILL = PatternFill(fill_type="solid", start_color="D9EAD3", end_color="D9EAD3")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="595959")
NORMAL_FONT = Font(name=FONT_NAME, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def build_ticker_sheet(wb, symbol, result, generated_at):
    ws = wb.create_sheet(title=symbol[:31])
    company_name = result.get('company_name', symbol)
    ws["A1"] = f"{company_name} ({symbol}) \u2014 Max Pain Detail"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"Expiration: {result.get('expiration','N/A')}    Last price: {result.get('price', 0.0):.2f}    "
                f"Generated: {generated_at}    (highlighted row = Max Pain strike)")
    ws["A2"].font = NOTE_FONT

    headers = ["Strike", "Call OI", "Put OI", "Call Pain", "Put Pain", "Total Pain"]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    rows = result.get("rows", [])
    first_data_row = header_row + 1
    last_data_row = header_row + len(rows)

    if not rows:
        return first_data_row, last_data_row

    for i, (k, c, p) in enumerate(rows):
        r = first_data_row + i
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=c)
        ws.cell(row=r, column=3, value=p)
        ws.cell(row=r, column=4, value=(
            f"=SUMPRODUCT((A{r}-$A${first_data_row}:$A${last_data_row})"
            f"*($A${first_data_row}:$A${last_data_row}<A{r})"
            f"*$B${first_data_row}:$B${last_data_row})"
        ))
        ws.cell(row=r, column=5, value=(
            f"=SUMPRODUCT(($A${first_data_row}:$A${last_data_row}-A{r})"
            f"*($A${first_data_row}:$A${last_data_row}>A{r})"
            f"*$C${first_data_row}:$C${last_data_row})"
        ))
        ws.cell(row=r, column=6, value=f"=D{r}+E{r}")
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = NORMAL_FONT

    rng = f"A{first_data_row}:F{last_data_row}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f"$F{first_data_row}=MIN($F${first_data_row}:$F${last_data_row})"],
                    fill=MAXPAIN_FILL),
    )

    for i, w in enumerate([10, 10, 10, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    chart = BarChart()
    chart.type = "col"
    chart.title = f"{symbol} Open Interest by Strike"
    chart.y_axis.title = "Open Interest"
    chart.x_axis.title = "Strike"
    chart.height, chart.width = 9, 20
    data = Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=last_data_row)
    cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"H{header_row}")

    return first_data_row, last_data_row


def build_dashboard(wb, ticker_results, sheet_ranges, generated_at, strike_range_pct, max_strikes):
    ws = wb.create_sheet(title="Dashboard", index=0)
    ws["A1"] = "Max Pain Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {generated_at}"
    ws["A2"].font = NORMAL_FONT
    ws["A3"] = (f"Source: Yahoo Finance (yfinance), EOD open interest. "
                f"Strike window: \u00b1{strike_range_pct:g}% of spot, capped at {max_strikes} strikes.")
    ws["A3"].font = NOTE_FONT

    headers = ["Ticker", "Company Name", "Status", "Last Price", "Expiration", "Max Pain Strike",
               "Distance to Max Pain (%)", "Total Call OI", "Total Put OI", "Put/Call Ratio", "Note"]
    header_row = 5
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    first_data_row = header_row + 1
    r = first_data_row
    for symbol, result in ticker_results.items():
        ws.cell(row=r, column=1, value=symbol)
        ws.cell(row=r, column=2, value=result.get("company_name", symbol))
        
        if result["status"] == "OK":
            sheet = symbol[:31]
            fdr, ldr = sheet_ranges[symbol]
            cell = ws.cell(row=r, column=3, value="OK")
            cell.fill = OK_FILL
            ws.cell(row=r, column=4, value=result["price"])
            ws.cell(row=r, column=5, value=result["expiration"])
            ws.cell(row=r, column=6, value=(
                f"=INDEX('{sheet}'!$A${fdr}:$A${ldr},"
                f"MATCH(MIN('{sheet}'!$F${fdr}:$F${ldr}),'{sheet}'!$F${fdr}:$F${ldr},0))"
            ))
            dist_cell = ws.cell(row=r, column=7, value=f"=(D{r}-F{r})/F{r}")
            dist_cell.number_format = "0.0%"
            ws.cell(row=r, column=8, value=f"=SUM('{sheet}'!$B${fdr}:$B${ldr})")
            ws.cell(row=r, column=9, value=f"=SUM('{sheet}'!$C${fdr}:$C${ldr})")
            ratio_cell = ws.cell(row=r, column=10, value=f'=IF(H{r}=0,"",I{r}/H{r})')
            ratio_cell.number_format = "0.00"
            ws.cell(row=r, column=11, value="")
        else:
            cell = ws.cell(row=r, column=3, value="FAILED")
            cell.fill = FAIL_FILL
            for col in range(4, 11):
                ws.cell(row=r, column=col, value="")
            ws.cell(row=r, column=11, value=result.get("error", ""))
        for col in range(1, 12):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).font = NORMAL_FONT
        r += 1
    last_data_row = r - 1

    for i, w in enumerate([10, 20, 10, 12, 13, 15, 22, 14, 14, 14, 55], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    chart = BarChart()
    chart.type = "col"
    chart.title = "Distance to Max Pain by Ticker (%)"
    chart.y_axis.title = "% from current price"
    chart.height, chart.width = 9, 20
    data = Reference(ws, min_col=7, min_row=header_row, max_row=last_data_row)
    cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{last_data_row + 3}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Batch max-pain fetch -> Excel dashboard")
    ap.add_argument("config", help="Path to JSON file listing tickers")
    ap.add_argument("--output", default="max_pain_dashboard.xlsx")
    ap.add_argument("--wait-until", default=None, help="HH:MM local time to wait for before starting, e.g. 22:30")
    ap.add_argument("--timezone", default="Europe/Berlin", help="IANA timezone for --wait-until (default auto-handles CET/CEST)")
    ap.add_argument("--max-retry-minutes", type=float, default=45.0, help="Keep retrying a ticker up to this long before marking it FAILED")
    ap.add_argument("--retry-interval-sec", type=float, default=120.0, help="Seconds between retries")
    ap.add_argument("--demo", action="store_true", help="Use synthetic data -- no network, no waiting, instant preview")
    args = ap.parse_args()

    cfg = load_config(args.config)

    if args.wait_until and not args.demo:
        sleep_until(args.wait_until, args.timezone)

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    results = {}
    for t in cfg["tickers"]:
        print(f"\n=== {t['symbol']} ===")
        results[t["symbol"]] = fetch_with_retry(
            t, cfg["strike_range_pct"], cfg["max_strikes"],
            args.max_retry_minutes, args.retry_interval_sec, args.demo,
        )

    wb = Workbook()
    wb.remove(wb.active)

    sheet_ranges = {}
    for symbol, result in results.items():
        if result["status"] == "OK":
            sheet_ranges[symbol] = build_ticker_sheet(wb, symbol, result, generated_at)

    build_dashboard(wb, results, sheet_ranges, generated_at, cfg["strike_range_pct"], cfg["max_strikes"])

    wb.save(args.output)
    print(f"\nSaved: {args.output}")


if __name__ == "__main__":
    main()
